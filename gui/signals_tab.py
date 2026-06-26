"""Mapped signals table — shows all discovered signal mappings."""
# ─────────────────────────────────────────────────────────────────────────────
# MÓDULO: Aba "Sinais Mapeados" (PyQt5)
#
# Este arquivo implementa a aba que mostra todos os sinais CAN descobertos pelo
# processo de discovery e oferece as ferramentas para transformar esse
# mapeamento no formato que o equipamento VIRLOC entende.
#
# Visão geral do que existe aqui:
#   - _MetadataDialog ...... janela que coleta os dados do veículo (modelo, ano,
#                            responsável, baudrate...) usados nos cabeçalhos de
#                            exportação.
#   - _CalibrationDialog ... assistente de calibração de 2 pontos para sinais
#                            proprietários, lendo o valor BRUTO ao vivo do CAN.
#   - funções _build_vs_string / _factor_to_ops / _build_mat_string / _mat_preview
#                            ... convertem um sinal descoberto para as strings
#                            VS e MAT do protocolo de configuração VIRLOC.
#   - SignalsTab ........... o widget principal da aba: a tabela de sinais, a
#                            exportação para Excel (formato VIRLOC), a geração
#                            do relatório PDF e o botão de calibração.
#
# IMPORTANTE: o programa é compilado em .exe e está em produção. Nada de lógica
# foi alterado aqui — apenas comentários explicativos foram adicionados.
# ─────────────────────────────────────────────────────────────────────────────

import json            # (importado para uso geral; serialização JSON)
import os              # (utilidades de caminho/sistema de arquivos)
from datetime import datetime   # carimbos de data/hora (data de mapeamento, duração)
from typing import Optional     # anotação de tipo para valores que podem ser None

from PyQt5.QtCore import Qt, pyqtSlot, QMarginsF, QTimer
from PyQt5.QtGui import QColor, QTextDocument, QPageLayout, QPageSize
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
    QGroupBox, QTextEdit, QMessageBox, QFrame,
    QDialog, QFormLayout, QLineEdit, QDialogButtonBox,
)

# Imports do próprio projeto:
from core.discovery import TESTS, CandidateSignal   # TESTS = catálogo de sinais; CandidateSignal = sinal candidato
from core.j1939 import decode_29bit_id              # decodifica um ID estendido de 29 bits em (PGN, PF, SA...)
from gui.styles import COLORS                        # paleta de cores da interface


# ── Nome amigável de cada sinal (coluna "INFORMAÇÕES") ──────────────────────
# Mapeia a CHAVE interna de cada sinal (ex: "rpm") para o nome legível em
# português que aparece na coluna "INFORMAÇÕES" da planilha exportada.
_SIGNAL_INFO_NAME = {
    # Motor / Funcionamento
    "rpm":             "RPM",
    "speed":           "Velocidade",
    "throttle":        "Pedal do acelerador",
    "clutch":          "Pedal de embreagem",
    "brake_pedal":     "Pedal do freio",
    "parking_brake":   "Freio estacionário",
    "gear":            "Câmbio",
    "coolant_temp":    "Temperatura arrefecimento",
    "oil_temp":        "Temperatura óleo",
    "oil_pressure":    "Pressão do óleo",
    "horimeter":       "Horímetro",
    "fuel_total":      "Consumo combustível acumulado",
    "trip_distance":   "Filtro de hodômetro",
    "odometer":        "Hodômetro",
    # Informações Gerais
    "ignition":        "Ignição",
    "air_conditioner": "Ar condicionado",
    "battery_voltage": "Voltagem da bateria",
    "fuel_level":      "Nível do tanque",
    "seatbelt":        "Cinto de segurança",
    "wiper":           "Limpador de parabrisa",
    "driver_door":     "Porta do motorista",
    "passenger_door":  "Porta do passageiro",
    "air_pressure":    "Pressão do ar",
}

# ── Unidade de medida de cada sinal ─────────────────────────────────────────
# Mapeia a CHAVE interna de cada sinal para sua unidade de engenharia. Usado na
# coluna "UNIDADE MEDIDA" do Excel e nos textos das fórmulas. "bool" indica um
# sinal liga/desliga; "estado" indica um sinal com múltiplos estados discretos.
_SIGNAL_UNIT = {
    "rpm":             "rpm",
    "speed":           "km/h",
    "throttle":        "%",
    "clutch":          "bool",
    "brake_pedal":     "%",
    "parking_brake":   "bool",
    "gear":            "marcha",
    "coolant_temp":    "°C",
    "oil_temp":        "°C",
    "oil_pressure":    "kPa",
    "horimeter":       "h",
    "fuel_total":      "L",
    "trip_distance":   "km",
    "odometer":        "km",
    "ignition":        "estado",
    "air_conditioner": "bool",
    "battery_voltage": "V",
    "fuel_level":      "%",
    "seatbelt":        "bool",
    "wiper":           "estado",
    "driver_door":     "bool",
    "passenger_door":  "bool",
    "air_pressure":    "kPa",
}


# ─────────────────────────────────────────────────────────────────────────────
class _MetadataDialog(QDialog):
    """Coleta as informações de cabeçalho (data, responsável, modelo, etc).

    É uma janela modal simples baseada em formulário (QFormLayout). Os valores
    digitados são reaproveitados tanto na exportação Excel quanto no relatório
    PDF — por isso esta classe é abstraída e chamada nos dois fluxos.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Dados do Mapeamento")
        self.setMinimumWidth(420)

        # QFormLayout organiza pares "rótulo: campo" em duas colunas alinhadas.
        layout = QFormLayout(self)
        layout.setSpacing(10)

        # Campos de entrada. Alguns já vêm preenchidos com valores padrão úteis:
        self.data        = QLineEdit(datetime.now().strftime("%d.%m.%Y"))  # data de hoje (dd.mm.aaaa)
        self.responsavel = QLineEdit()                                     # quem fez o mapeamento
        self.baudrate    = QLineEdit("250")                               # baudrate típico CAN (250 kbps)
        self.modelo      = QLineEdit()                                    # modelo do veículo
        self.fabricante  = QLineEdit()                                    # montadora
        self.ano         = QLineEdit(str(datetime.now().year))            # ano (padrão: ano atual)

        # Dicas de preenchimento exibidas em cinza dentro dos campos vazios.
        self.responsavel.setPlaceholderText("Ex: Fernando/Williams")
        self.modelo.setPlaceholderText("Ex: 17230OD")
        self.fabricante.setPlaceholderText("Ex: Volkswagen")

        # Adiciona cada campo ao formulário com seu rótulo correspondente.
        layout.addRow("Data de mapeamento:", self.data)
        layout.addRow("Responsável:",        self.responsavel)
        layout.addRow("Baudrate (kbps):",    self.baudrate)
        layout.addRow("Modelo:",             self.modelo)
        layout.addRow("Fabricante:",         self.fabricante)
        layout.addRow("Ano:",                self.ano)

        # Botões padrão OK/Cancelar. accept() fecha com QDialog.Accepted,
        # reject() fecha com QDialog.Rejected (verificado por quem chamou).
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def values(self) -> dict:
        """Retorna os dados digitados como dicionário, já com espaços aparados.

        Chamado pelo código de exportação depois que o usuário clica em OK.
        """
        return {
            "data":        self.data.text().strip(),
            "responsavel": self.responsavel.text().strip(),
            "baudrate":    self.baudrate.text().strip(),
            "modelo":      self.modelo.text().strip(),
            "fabricante":  self.fabricante.text().strip(),
            "ano":         self.ano.text().strip(),
        }


# ─────────────────────────────────────────────────────────────────────────────
class _CalibrationDialog(QDialog):
    """
    Assistente de calibração de 2 pontos para sinais proprietários.

    Lê AO VIVO o valor bruto do byte do sinal selecionado. O usuário captura
    2 pontos (informando o valor REAL mostrado no painel em cada um) e o
    programa calcula fator e offset exatos por regressão linear:

        factor = (real2 - real1) / (raw2 - raw1)
        offset =  real1 - raw1 * factor
    """

    def __init__(self, bus, sig: CandidateSignal, signal_name: str, parent=None):
        """Inicializa o assistente de calibração.

        Parâmetros:
            bus         -- referência ao CANBus conectado, de onde o valor bruto
                           ao vivo será lido (via listener).
            sig         -- o CandidateSignal a calibrar; define qual CAN ID e
                           quais bytes observar.
            signal_name -- nome legível do sinal, usado no título da janela.

        THREAD SAFETY: o CAN roda em uma thread própria (de I/O), enquanto a
        interface Qt roda na thread principal. O listener (_on_msg) é chamado
        pela thread do CAN e SÓ grava um inteiro em self._live_raw — operação
        atômica garantida pelo GIL do Python. A leitura/exibição desse valor é
        feita pela thread principal através de um QTimer (_refresh_live). Assim
        evitamos tocar em widgets Qt fora da thread principal.
        """
        super().__init__(parent)
        self._bus  = bus     # barramento CAN (para add_listener/remove_listener)
        self._sig  = sig     # sinal sendo calibrado (define CAN ID + bytes)
        self.setWindowTitle(f"Calibração — {signal_name}")
        self.setMinimumWidth(460)

        self._live_raw = None         # último valor bruto lido (escrito pela thread CAN)
        self._p1_raw = None           # valor bruto capturado no Ponto 1
        self._p2_raw = None           # valor bruto capturado no Ponto 2
        self._result = (None, None)   # (fator, offset) calculado; lido por quem abriu o diálogo

        self._setup_ui(signal_name)   # monta todos os widgets

        # Registra o listener no bus para receber TODAS as mensagens CAN e
        # filtrar (dentro de _on_msg) apenas as do CAN ID deste sinal.
        self._bus.add_listener(self._on_msg)
        # Timer que dispara periodicamente NA THREAD PRINCIPAL para atualizar o
        # rótulo do valor ao vivo — única thread autorizada a mexer em widgets.
        self._timer = QTimer(self)
        self._timer.setInterval(120)               # a cada 120 ms (~8 Hz)
        self._timer.timeout.connect(self._refresh_live)
        self._timer.start()

    def _setup_ui(self, signal_name: str):
        """Constrói todos os widgets da janela de calibração."""
        from PyQt5.QtWidgets import QGridLayout
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # Cabeçalho informativo: ID em hex (8 dígitos se 29-bit, 4 se 11-bit) e
        # o(s) byte(s) que serão lidos (faixa "início-fim" se for mais de 1 byte).
        id_str = (f"0x{self._sig.can_id:08X}" if self._sig.is_extended
                  else f"0x{self._sig.can_id:04X}")
        byte_s = (f"{self._sig.byte_index}" if self._sig.length_bytes == 1
                  else f"{self._sig.byte_index}-{self._sig.byte_index + self._sig.length_bytes - 1}")
        info = QLabel(f"<b>{signal_name}</b><br>"
                      f"CAN ID: {id_str} &nbsp; | &nbsp; Byte(s): {byte_s} &nbsp; | &nbsp; "
                      f"{self._sig.length_bytes} byte(s) {self._sig.byte_order}")
        info.setStyleSheet(f"color: {COLORS['text']}; font-size: 12px;")
        layout.addWidget(info)

        # Valor bruto ao vivo — rótulo grande e destacado, atualizado pelo timer.
        self._lbl_live = QLabel("Valor bruto atual: —")
        self._lbl_live.setStyleSheet(
            f"color: {COLORS['accent']}; font-size: 18px; font-weight: bold; "
            f"padding: 8px; background: #0d0d1a; border-radius: 4px;")
        self._lbl_live.setAlignment(Qt.AlignCenter)
        layout.addWidget(self._lbl_live)

        # Instrução
        tip = QLabel(
            "1. Coloque o sinal em um valor conhecido (ex: 20 km/h no painel)\n"
            "2. Digite o valor REAL e clique 'Capturar Ponto 1'\n"
            "3. Mude para outro valor (ex: 60 km/h) e capture o Ponto 2\n"
            "4. Clique 'Calcular' — o fator/offset serão calculados.")
        tip.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 11px;")
        layout.addWidget(tip)

        # Grade de captura dos 2 pontos. Cada linha tem: rótulo, campo do valor
        # REAL (lido no painel), botão "Capturar" e rótulo do valor bruto pego.
        grid = QGridLayout()
        grid.setSpacing(8)
        # Ponto 1
        grid.addWidget(QLabel("Ponto 1 — valor real:"), 0, 0)
        self._inp_real1 = QLineEdit()
        self._inp_real1.setPlaceholderText("ex: 20")
        grid.addWidget(self._inp_real1, 0, 1)
        self._btn_cap1 = QPushButton("Capturar Ponto 1")
        self._btn_cap1.clicked.connect(self._capture_p1)
        grid.addWidget(self._btn_cap1, 0, 2)
        self._lbl_p1 = QLabel("bruto: —")
        self._lbl_p1.setStyleSheet(f"color: {COLORS['text_muted']};")
        grid.addWidget(self._lbl_p1, 0, 3)
        # Ponto 2
        grid.addWidget(QLabel("Ponto 2 — valor real:"), 1, 0)
        self._inp_real2 = QLineEdit()
        self._inp_real2.setPlaceholderText("ex: 60")
        grid.addWidget(self._inp_real2, 1, 1)
        self._btn_cap2 = QPushButton("Capturar Ponto 2")
        self._btn_cap2.clicked.connect(self._capture_p2)
        grid.addWidget(self._btn_cap2, 1, 2)
        self._lbl_p2 = QLabel("bruto: —")
        self._lbl_p2.setStyleSheet(f"color: {COLORS['text_muted']};")
        grid.addWidget(self._lbl_p2, 1, 3)
        layout.addLayout(grid)

        # Resultado — exibe a fórmula calculada (decimal e VIRLOC) em verde.
        self._lbl_result = QLabel("")
        self._lbl_result.setStyleSheet(
            f"color: {COLORS['success']}; font-size: 13px; font-weight: bold;")
        self._lbl_result.setWordWrap(True)
        layout.addWidget(self._lbl_result)

        # Linha de botões: Calcular / Aplicar / Cancelar.
        btn_row = QHBoxLayout()
        self._btn_calc = QPushButton("🧮  Calcular")
        self._btn_calc.setObjectName("btn_success")
        self._btn_calc.clicked.connect(self._calculate)   # dispara a regressão de 2 pontos
        btn_row.addWidget(self._btn_calc)
        self._btn_apply = QPushButton("✓  Aplicar")
        self._btn_apply.clicked.connect(self.accept)      # fecha aceitando; resultado é lido depois
        self._btn_apply.setEnabled(False)                 # só habilita após um cálculo bem-sucedido
        btn_row.addWidget(self._btn_apply)
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _extract_raw(self, msg) -> Optional[int]:
        """Extrai o valor bruto inteiro do(s) byte(s) alvo de uma mensagem CAN.

        Lê 'length_bytes' bytes a partir de 'byte_index'. Para 1 byte retorna o
        próprio valor; para vários bytes monta o inteiro respeitando o
        byte_order (little/big endian). Retorna None se a mensagem for curta
        demais para conter os bytes esperados.
        """
        bi  = self._sig.byte_index      # índice do primeiro byte (0-based)
        ln  = self._sig.length_bytes    # quantos bytes compõem o valor
        data = msg.data
        if bi + ln > len(data):         # mensagem não tem bytes suficientes
            return None
        if ln == 1:
            return data[bi]             # caso simples: 1 byte
        chunk = bytes(data[bi:bi + ln])
        # Define a ordem de bytes para reconstruir o inteiro de múltiplos bytes.
        order = 'little' if (self._sig.byte_order or 'little') == 'little' else 'big'
        return int.from_bytes(chunk, order)

    def _on_msg(self, msg):
        """Listener chamado PELA THREAD DO CAN a cada mensagem recebida.

        Filtra pelo CAN ID do sinal e, se bater, extrai o valor bruto e o grava
        em self._live_raw. A escrita de um único inteiro é atômica sob o GIL,
        então NÃO precisamos de lock aqui. Nunca toca em widgets Qt (isso fica
        a cargo de _refresh_live, na thread principal).
        """
        if msg.can_id == self._sig.can_id:
            raw = self._extract_raw(msg)
            if raw is not None:
                self._live_raw = raw

    @pyqtSlot()
    def _refresh_live(self):
        """Atualiza o rótulo do valor ao vivo (chamado pelo QTimer, thread principal)."""
        if self._live_raw is not None:
            self._lbl_live.setText(
                f"Valor bruto atual: {self._live_raw}  (0x{self._live_raw:X})")

    @pyqtSlot()
    def _capture_p1(self):
        """Congela o valor bruto atual como Ponto 1 da calibração."""
        if self._live_raw is None:
            QMessageBox.warning(self, "Sem dados",
                                "Nenhum valor recebido ainda para este CAN ID.")
            return
        self._p1_raw = self._live_raw   # memoriza o bruto deste instante
        self._lbl_p1.setText(f"bruto: {self._p1_raw}")

    @pyqtSlot()
    def _capture_p2(self):
        """Congela o valor bruto atual como Ponto 2 da calibração."""
        if self._live_raw is None:
            QMessageBox.warning(self, "Sem dados",
                                "Nenhum valor recebido ainda para este CAN ID.")
            return
        self._p2_raw = self._live_raw   # memoriza o bruto deste instante
        self._lbl_p2.setText(f"bruto: {self._p2_raw}")

    @pyqtSlot()
    def _calculate(self):
        """Calcula fator e offset por regressão linear de 2 pontos.

        A reta que passa pelos dois pontos (raw1, real1) e (raw2, real2) é:
            real = raw * fator + offset
        onde:
            fator  = (real2 - real1) / (raw2 - raw1)
            offset =  real1 - raw1 * fator

        O resultado é apresentado em DUAS formas:
          1) Decimal — exata, para conferência humana.
          2) VIRLOC (inteiros) — o fator vira operações inteiras (x/) e o offset
             é arredondado para inteiro, pois o VIRLOC só aceita inteiros.
        Também mostra uma verificação aplicando a fórmula inteira nos 2 brutos
        capturados, para o usuário ver o erro de arredondamento, se houver.
        """
        # Aceita vírgula como separador decimal (digitação em pt-BR).
        try:
            real1 = float(self._inp_real1.text().replace(",", "."))
            real2 = float(self._inp_real2.text().replace(",", "."))
        except ValueError:
            QMessageBox.warning(self, "Valores inválidos",
                                "Digite os valores reais (numéricos) dos 2 pontos.")
            return
        # Precisa dos 2 brutos capturados.
        if self._p1_raw is None or self._p2_raw is None:
            QMessageBox.warning(self, "Pontos faltando",
                                "Capture os 2 pontos antes de calcular.")
            return
        # Se os brutos forem idênticos, a inclinação seria divisão por zero.
        if self._p1_raw == self._p2_raw:
            QMessageBox.warning(self, "Pontos iguais",
                                "Os valores brutos dos 2 pontos são iguais.\n"
                                "Use valores reais bem diferentes (ex: 20 e 60).")
            return
        from fractions import Fraction
        # Calcula o fator como FRAÇÃO EXATA dos deltas. Fraction(str(...)) evita
        # o ruído binário de float (ex: 0.1 não é exatamente 0.1 em binário).
        # limit_denominator(10000) reduz à melhor fração com denominador <= 10000,
        # que é a forma compatível com as operações inteiras do VIRLOC.
        dreal = Fraction(str(real2)) - Fraction(str(real1))   # Δreal
        draw  = self._p2_raw - self._p1_raw                   # Δraw (inteiro)
        frac  = (dreal / draw).limit_denominator(10000)       # fator = Δreal/Δraw
        # Offset exato pela reta, depois arredondado para inteiro (exigência VIRLOC).
        offset_exact = float(real1) - self._p1_raw * float(frac)
        off_int = int(round(offset_exact))

        factor = float(frac)             # fator em precisão total (não arredonda p/ 6 dígitos)
        self._result = (factor, float(off_int))   # par devolvido a quem abriu o diálogo

        # Prévia da forma inteira VIRLOC (ex: "raw x2/5 +3").
        mat_int = _mat_preview(factor, float(off_int))
        # Verificação: aplica a fórmula INTEIRA (fração + offset inteiro) nos
        # brutos capturados para mostrar o resultado reconstruído ao usuário.
        chk1 = float(self._p1_raw * frac + off_int)
        chk2 = float(self._p2_raw * frac + off_int)

        sign = "+" if offset_exact >= 0 else "-"   # sinal do offset para o texto decimal
        self._lbl_result.setText(
            f"✓ Decimal:  valor = raw × {factor:.6g} {sign} {abs(offset_exact):.4g}\n"
            f"✓ VIRLOC (inteiros):  {mat_int}\n"
            f"   Confere: raw {self._p1_raw} → {chk1:.1f}  |  "
            f"raw {self._p2_raw} → {chk2:.1f}  (real: {real1:g} / {real2:g})")
        self._btn_apply.setEnabled(True)   # libera o botão "Aplicar" agora que há resultado

    def result_factor_offset(self):
        """Retorna a tupla (fator, offset) calculada — lida após accept()."""
        return self._result

    # As três funções abaixo (closeEvent/reject/accept) garantem que o listener
    # do CAN e o timer sejam SEMPRE desligados ao fechar a janela, qualquer que
    # seja o caminho de saída (X, Cancelar ou Aplicar). Deixar o listener
    # pendurado causaria callbacks numa janela já destruída. O try/except cobre
    # o caso de o bus já ter sido desconectado.
    def closeEvent(self, event):
        """Fechamento pelo X da janela: para o timer e remove o listener."""
        try:
            self._timer.stop()
            self._bus.remove_listener(self._on_msg)
        except Exception:
            pass
        event.accept()

    def reject(self):
        """Saída por Cancelar: limpa recursos antes de rejeitar o diálogo."""
        try:
            self._timer.stop()
            self._bus.remove_listener(self._on_msg)
        except Exception:
            pass
        super().reject()

    def accept(self):
        """Saída por Aplicar: limpa recursos antes de aceitar o diálogo."""
        try:
            self._timer.stop()
            self._bus.remove_listener(self._on_msg)
        except Exception:
            pass
        super().accept()


def _build_vs_string(vs_filter_num: int, ct_num: int, sig: CandidateSignal) -> str:
    """
    Monta string VS conforme documentação VIRLOC:
    VS19xx,ID,SA,IDbits,CT,byte_start,length,mask,mode,byte_order,multiplier

    O VIRLOC é o equipamento de telemetria de destino. Um "filtro VS" diz a ele
    QUAIS bytes de QUAL mensagem CAN capturar e onde guardar o resultado. Cada
    campo da string, na ordem em que aparece:

      VS19xx ...... cabeçalho + número do filtro (00..24). São 25 filtros
                    possíveis: VS1900 a VS1924.
      ID .......... identificador da mensagem a capturar. Para J1939 (29-bit) é
                    o PGN extraído; para proprietário (11-bit) é o próprio CAN
                    ID (0x000..0x7FF). Formatado com 5 dígitos decimais.
      SA .......... Source Address (endereço de origem J1939, 3 dígitos). Só faz
                    sentido em 29-bit; em 11-bit fica 000.
      IDbits ...... 29 (estendido/J1939) ou 11 (padrão/proprietário).
      CT .......... número do "Contador/Canal" de destino no VIRLOC (01..96)
                    onde o valor capturado será armazenado.
      byte_start .. posição do primeiro byte dentro do payload, 1-INDEXADO
                    (1..8). Atenção: internamente byte_index é 0-based, por isso
                    somamos 1.
      length ...... quantos bytes ler (1..4).
      mask ........ máscara de bits aplicada ao valor, 8 dígitos hex (4 bytes).
                    Aqui é sempre "todos os bits do tamanho" (ex: 1 byte = 000000FF).
      mode ........ 0 = copia o valor / 1 = acumula (somar a cada leitura).
      byte_order .. 0 = INVERTE os bytes (J1939 é little-endian) /
                    1 = não inverte (big-endian / Motorola).
      multiplier .. multiplicador fixo "01000" (x1000/1000) que preserva
                    precisão inteira no transporte do valor.

    Parâmetros:
      - vs_filter_num: 00..24 (índice do filtro S19)
      - ct_num:        01..96 (CT destino do valor capturado)
      - byte_start:    1..8   (1-indexed!)
      - length:        1..4   bytes
      - mode:          0 = copia / 1 = acumula
      - byte_order:    0 = inverte (J1939 little-endian) / 1 = não inverte
      - multiplier:    1000   (padrão: x1000/1000 — preserva precisão inteira)
    """
    # Source Address: extraído do ID de 29 bits; permanece 0 para 11-bit.
    sa = 0
    if sig.is_extended:
        _, _, sa = decode_29bit_id(sig.can_id)
    sa_str     = f"{sa:03d}"                  # SA com 3 dígitos (ex: "000")
    id_bits    = 29 if sig.is_extended else 11   # 29-bit vs 11-bit

    # Campo "ID" do filtro: PGN (J1939) ou CAN ID bruto (proprietário 11-bit)
    if sig.is_extended and sig.pgn:
        id_field = sig.pgn                    # J1939 com PGN conhecido → usa o PGN
    else:
        # 11-bit ou 29-bit sem PGN identificado → usa o próprio CAN ID.
        # A máscara de 29 bits (0x1FFFFFFF) cobre os dois casos sem alterar
        # um ID de 11 bits (que já cabe nela).
        id_field = sig.can_id & 0x1FFFFFFF  # 29-bit mask, suporta ambos os casos

    byte_start = sig.byte_index + 1                # converte 0-based → 1-based (exigência VIRLOC)
    length     = max(1, min(sig.length_bytes, 4))  # limita ao intervalo válido 1..4
    # Máscara = todos os bits ocupados pelos 'length' bytes, ex:
    #   length 1 → 0xFF      → "000000FF"
    #   length 2 → 0xFFFF    → "0000FFFF"
    # (1 << (length*8)) - 1 gera exatamente esse valor; :08X formata 8 hex.
    mask       = f"{(1 << (length * 8)) - 1:08X}"   # sempre 8 hex (4 pares)
    mode       = 0                                  # 0 = copia o valor (não acumula)
    # J1939 little-endian → INVERTE bytes (=0); big-endian/Motorola → não inverte (=1)
    byte_order = 0 if (sig.byte_order or "little") == "little" else 1
    multiplier = "01000"                            # multiplicador padrão fixo

    # Concatena tudo na ordem documentada. f-strings com :02d/:05d/:03d apenas
    # formatam a largura/zero-padding de cada campo — NÃO alteram os valores.
    return (f"VS19{vs_filter_num:02d},{id_field:05d},{sa_str},{id_bits},"
            f"{ct_num:02d},{byte_start},{length},{mask},"
            f"{mode},{byte_order},{multiplier}")


def _factor_to_ops(factor: float, max_den: int = 10000) -> list[str]:
    """
    Converte um fator decimal em operações INTEIRAS MAT (xN, /N), pois o
    VIRLOC só aceita inteiros.

    Usa aproximação racional (fração) — qualquer decimal vira N/M com o
    menor denominador possível dentro da tolerância.

    POR QUE FRAÇÕES? O VIRLOC não tem ponto flutuante — ele só sabe multiplicar
    e dividir por inteiros. Logo, um fator decimal como 0.4 precisa virar uma
    sequência de operações inteiras. Fraction(factor).limit_denominator(max_den)
    encontra a fração N/M com o MENOR denominador (<= max_den) que melhor
    aproxima o decimal. Aplicar "xN /M" ao valor bruto reproduz o fator decimal
    usando apenas inteiros, sem perder precisão prática.

    Exemplos:
        0.125      → /8          (porque 0.125 = 1/8)
        0.4        → x2/5        (porque 0.4 = 2/5: multiplica por 2, divide por 5)
        0.222469   → x200/899   (melhor fração dentro da tolerância — aproximação ótima)
        2.0        → x2          (inteiro puro, só multiplica)
    """
    from fractions import Fraction
    # Fator nulo ou 1.0 não altera nada → nenhuma operação necessária.
    if not factor or factor == 1.0:
        return []
    # Melhor fração N/M com denominador <= max_den que aproxima o decimal.
    frac = Fraction(factor).limit_denominator(max_den)
    num, den = frac.numerator, frac.denominator

    if den == 1:
        # Denominador 1 → fator inteiro puro → basta multiplicar (xN).
        # Se num também for 1, o fator era 1 → nenhuma operação.
        return [f"x{num}"] if num != 1 else []
    if num == 1:
        # Numerador 1 → fração 1/M → basta dividir (/M).
        return [f"/{den}"]
    # Caso geral N/M: multiplica por N e depois divide por M (ex: "x2/5").
    return [f"x{num}/{den}"]


def _factor_to_ops_str(factor: float) -> str:
    """Retorna a representação textual das operações (ex: 'x200/899').

    Concatena a lista de operações de _factor_to_ops numa única string. Se não
    houver operação (fator 1 ou nulo), devolve "x1" como neutro legível.
    """
    ops = _factor_to_ops(factor)
    return "".join(ops) if ops else "x1"


def _mat_preview(factor: float, offset: float) -> str:
    """
    Monta uma prévia legível das operações inteiras VIRLOC para um
    par (fator, offset). Ex: 'raw x200/899 -902'

    Usada na janela de calibração para mostrar ao usuário, em texto, como o
    valor bruto será transformado: começa em "raw", aplica as operações do
    fator e, por fim, soma/subtrai o offset (arredondado para inteiro).
    """
    parts = ["raw"]                    # ponto de partida: o valor bruto
    ops = _factor_to_ops(factor)       # operações inteiras do fator (ex: x2/5)
    parts.extend(ops)
    off_int = int(round(offset or 0))  # offset arredondado para inteiro (VIRLOC)
    if off_int < 0:
        parts.append(f"{off_int}")     # já vem com o sinal "-", ex: -902
    elif off_int > 0:
        parts.append(f"+{off_int}")    # offset positivo recebe "+" explícito
    return " ".join(parts)


def _build_mat_string(vs_filter_num: int, sig: CandidateSignal) -> str:
    """
    Monta string MAT conforme documentação VIRLOC:
    VS19xx_MAT,operações,...

    A string MAT (de "matemática") acompanha o filtro VS de mesmo número e diz
    ao VIRLOC como CONVERTER o valor bruto capturado em valor de engenharia. É
    uma lista de operações INTEIRAS aplicadas em sequência, da esquerda para a
    direita, sem precedência matemática (não é "fator depois soma" por regra de
    PEMDAS — é literalmente uma pilha de passos):
        VS19xx_MAT,x2,/5,+40   →  valor = ((raw * 2) / 5) + 40

    Operadores aceitos: x (mult), / (div), + (soma), - (sub), L (mín), H (máx).
    Apenas valores INTEIROS são permitidos — por isso o fator decimal é
    convertido em x/ via _factor_to_ops e o offset é arredondado para inteiro.
    """
    ops: list[str] = []

    # 1) Offset (soma/subtração) — aplica ANTES do fator? Documentação aplica
    #    da esquerda para direita SEM precedência matemática; típico SAE J1939
    #    é (raw * factor) + offset, então o offset entra ao final:
    # Fator de escala do sinal. Se for 1.0 não gera operação (multiplicar por 1
    # é inócuo). Caso contrário vira as operações inteiras x/ correspondentes.
    factor = sig.formula_factor or 1.0
    if factor != 1.0:
        ops.extend(_factor_to_ops(factor))

    # Offset aditivo, arredondado para inteiro e colocado APÓS o fator (ordem
    # raw*fator+offset). Negativo já traz o "-"; positivo recebe "+" explícito.
    offset = sig.formula_offset or 0.0
    if offset:
        off_int = int(round(offset))
        if off_int < 0:
            ops.append(f"{off_int}")        # ex: -125
        elif off_int > 0:
            ops.append(f"+{off_int}")       # ex: +40

    head = f"VS19{vs_filter_num:02d}_MAT"   # cabeçalho ligado ao filtro VS de mesmo número
    if not ops:
        return head                          # sem operações → só o cabeçalho (valor = raw)
    return head + "," + ",".join(ops)        # cabeçalho + operações separadas por vírgula


class SignalsTab(QWidget):
    """Aba principal "Sinais Mapeados".

    Mostra, em tabela, o melhor candidato de cada sinal descoberto e oferece as
    ações: exportar Excel (formato VIRLOC), gerar relatório PDF, calibrar um
    sinal selecionado e limpar a sessão. Mantém dois dicionários:
      - _mappings ......... apenas o MELHOR candidato por sinal (o que vai na
                            tabela e nas exportações).
      - _all_candidates ... todos os candidatos por sinal, para listar as
                            alternativas menos prováveis no relatório PDF.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        # Melhor candidato por sinal (chave do sinal → CandidateSignal).
        self._mappings: dict[str, CandidateSignal] = {}
        # Guarda TODOS os candidatos por sinal (top 1 + alternativos)
        # para que o relatório PDF possa exibir as PGNs menos prováveis.
        self._all_candidates: dict[str, list[CandidateSignal]] = {}
        self._session_start: Optional[datetime] = None   # início da sessão (1ª descoberta)
        self._bus = None                     # referência ao CANBus (p/ calibração)
        self._row_keys: list[str] = []       # mapeia índice de linha -> chave do sinal
        self._setup_ui()                     # monta a interface

    def set_bus(self, bus):
        """Recebe a referência do CANBus para permitir calibração ao vivo."""
        self._bus = bus

    def _setup_ui(self):
        """Monta o layout: cabeçalho com botões, linha de resumo e a tabela."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        # Header — título à esquerda e botões de ação à direita.
        hdr = QHBoxLayout()
        lbl = QLabel("Mapeamento de Sinais")
        lbl.setObjectName("label_title")
        hdr.addWidget(lbl)
        hdr.addStretch()

        self._btn_excel = QPushButton("📊  Exportar Excel")
        self._btn_excel.setObjectName("btn_success")
        self._btn_excel.clicked.connect(self._export_excel)
        hdr.addWidget(self._btn_excel)

        self._btn_pdf = QPushButton("📄  Relatório PDF")
        self._btn_pdf.clicked.connect(self._export_pdf_report)
        hdr.addWidget(self._btn_pdf)

        self._btn_calib = QPushButton("📐  Calibrar Sinal")
        self._btn_calib.setToolTip(
            "Calibra fator/offset do sinal SELECIONADO comparando o valor bruto\n"
            "do CAN com o valor real do painel. Ideal para protocolos proprietários.")
        self._btn_calib.clicked.connect(self._calibrate_selected)
        hdr.addWidget(self._btn_calib)

        self._btn_clear = QPushButton("🗑  Limpar")
        self._btn_clear.setObjectName("btn_danger")
        self._btn_clear.clicked.connect(self._clear)
        hdr.addWidget(self._btn_clear)
        layout.addLayout(hdr)

        # Summary row
        summary_row = QHBoxLayout()
        self._lbl_summary = QLabel("Nenhum sinal mapeado ainda.")
        self._lbl_summary.setStyleSheet(f"color: {COLORS['text_muted']};")
        summary_row.addWidget(self._lbl_summary)
        layout.addLayout(summary_row)

        # Tabela principal — 8 colunas, 0 linhas iniciais (preenchida depois).
        self._table = QTableWidget(0, 8)
        self._table.setHorizontalHeaderLabels([
            "Sinal", "CAN ID", "Tipo", "PGN", "Nome PGN", "Byte(s)", "Fórmula", "Confiança"
        ])
        # Modos de redimensionamento de cada coluna: a maioria se ajusta ao
        # conteúdo; a coluna 6 ("Fórmula") estica para ocupar o espaço restante.
        hdr2 = self._table.horizontalHeader()
        hdr2.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        hdr2.setSectionResizeMode(6, QHeaderView.Stretch)
        hdr2.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self._table.verticalHeader().setVisible(False)          # oculta numeração lateral de linhas
        self._table.setSelectionBehavior(QTableWidget.SelectRows)  # clicar seleciona a linha inteira
        layout.addWidget(self._table)

    # ── Public API ────────────────────────────────────────────────────────────

    def update_from_discovery(self, results_dict: dict[str, list[CandidateSignal]]):
        """Chamado quando a aba de discovery produz novos resultados.

        results_dict mapeia cada chave de sinal para uma LISTA de candidatos já
        ordenada por confiança (o primeiro é o melhor). Guarda todos para o
        relatório e promove o melhor à tabela se a confiança for suficiente.
        """
        # Marca o início da sessão na primeira descoberta bem-sucedida
        if self._session_start is None and any(c for c in results_dict.values()):
            self._session_start = datetime.now()
        # Confiança mínima para considerar um sinal "mapeado" automaticamente.
        # Abaixo disso o melhor candidato é provavelmente ruído (ex: 8% num byte
        # errado) e não deve poluir a tabela de mapeamento. Ainda assim é
        # guardado em _all_candidates para aparecer no relatório PDF.
        MIN_CONFIDENCE = 0.25
        for key, candidates in results_dict.items():
            if candidates:
                # Sempre guarda TODOS os candidatos (cópia da lista) para que o
                # relatório PDF possa exibir as alternativas menos prováveis.
                self._all_candidates[key] = list(candidates)
                # Só promove o melhor candidato à tabela se passar do limiar.
                if candidates[0].confidence >= MIN_CONFIDENCE:
                    self._mappings[key] = candidates[0]
        self._rebuild_table()   # redesenha a tabela com o estado atualizado

    # ── Table ─────────────────────────────────────────────────────────────────

    def _rebuild_table(self):
        """Redesenha a tabela inteira a partir de self._mappings.

        Também recria self._row_keys, de modo que o índice de cada linha
        corresponde à posição da chave nessa lista — isso permite descobrir
        qual sinal está selecionado ao calibrar.
        """
        self._table.setRowCount(0)   # limpa a tabela
        self._row_keys = []
        for key, sig in self._mappings.items():
            self._row_keys.append(key)   # índice da linha == posição nesta lista
            test = TESTS.get(key)
            signal_name = test.name if test else key

            row = self._table.rowCount()
            self._table.insertRow(row)

            # Formata cada coluna para exibição:
            id_str = f"0x{sig.can_id:08X}" if sig.is_extended else f"0x{sig.can_id:04X}"  # ID em hex
            type_str = "J1939 (29b)" if sig.is_extended else "Std (11b)"                   # tipo de quadro
            pgn_str = str(sig.pgn) if sig.pgn else "—"                                     # PGN ou travessão
            pgn_name = sig.pgn_name or "—"                                                 # nome do PGN
            byte_str = f"{sig.byte_index}"                                                 # byte(s) — aqui 0-based
            if sig.length_bytes > 1:
                byte_str += f"–{sig.byte_index + sig.length_bytes - 1}"   # faixa "início–fim" se >1 byte
            conf_str = f"{sig.confidence * 100:.0f}%"                                      # confiança em %
            # Cor da confiança: verde >80%, amarelo >50%, vermelho abaixo disso.
            conf_color = (COLORS['success'] if sig.confidence > 0.8 else
                          COLORS['warning'] if sig.confidence > 0.5 else COLORS['error'])

            # Texto e alinhamento de cada uma das 8 colunas, na ordem do cabeçalho.
            values = [signal_name, id_str, type_str, pgn_str, pgn_name,
                      byte_str, sig.formula_str, conf_str]
            aligns = [Qt.AlignLeft, Qt.AlignCenter, Qt.AlignCenter, Qt.AlignCenter,
                      Qt.AlignLeft, Qt.AlignCenter, Qt.AlignLeft, Qt.AlignCenter]

            for col, (text, align) in enumerate(zip(values, aligns)):
                item = QTableWidgetItem(text)
                item.setTextAlignment(align | Qt.AlignVCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)   # célula somente leitura
                if col == 7:
                    item.setForeground(QColor(conf_color))         # pinta a coluna de confiança
                self._table.setItem(row, col, item)

        # Linha de resumo: "X/Y sinais mapeados" + status (completo/pendentes).
        total = len(TESTS)             # total de sinais no catálogo
        found = len(self._mappings)    # quantos foram mapeados
        self._lbl_summary.setText(
            f"✅ {found}/{total} sinais mapeados  |  "
            f"{'Mapeamento completo!' if found == total else f'{total-found} sinal(is) pendente(s)'}"
        )
        self._lbl_summary.setStyleSheet(
            f"color: {COLORS['success']};" if found == total else f"color: {COLORS['warning']};"
        )

    # ── Actions ───────────────────────────────────────────────────────────────

    @pyqtSlot()
    def _export_excel(self):
        """Exporta o mapeamento como planilha Excel com cabeçalho de metadados.

        Usa a biblioteca openpyxl para gerar um .xlsx no formato VIRLOC. A
        planilha tem: (1) um cabeçalho de metadados do veículo nas primeiras
        linhas, (2) o cabeçalho da tabela na linha 4 e (3) uma linha por sinal,
        cada uma já com as strings VS e MAT prontas para o equipamento. Linhas
        alternam de cor (zebra) para legibilidade.
        """
        if not self._mappings:
            QMessageBox.warning(self, "Nada para exportar",
                                "Nenhum sinal mapeado ainda.")
            return

        # 1. Coleta os metadados
        dlg = _MetadataDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return
        meta = dlg.values()

        # 2. Pergunta o local de salvamento
        modelo_clean = meta["modelo"].replace(" ", "_") or "veiculo"
        default_name = f"mapeamento_{modelo_clean}_{meta['ano']}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar Excel", default_name, "Excel (*.xlsx)"
        )
        if not path:
            return

        # 3. Gera o arquivo
        # openpyxl é dependência opcional; importada aqui para dar mensagem
        # amigável caso não esteja instalada (em vez de quebrar o programa).
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self, "openpyxl não instalado",
                "É necessário instalar o módulo 'openpyxl' para exportar Excel.\n\n"
                "Execute no CMD:  pip install openpyxl"
            )
            return

        try:
            wb = Workbook()          # cria a pasta de trabalho
            ws = wb.active           # usa a primeira planilha
            ws.title = "Mapeamento"

            # Estilos reutilizáveis (definidos uma vez e aplicados a várias células):
            header_fill = PatternFill("solid", fgColor="C0C0C0")   # fundo cinza dos cabeçalhos
            row_fill_a  = PatternFill("solid", fgColor="FFFFFF")    # zebra: linha branca
            row_fill_b  = PatternFill("solid", fgColor="F2F2F2")    # zebra: linha cinza-claro
            bold        = Font(bold=True)
            thin        = Side(border_style="thin", color="808080")
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)
            center      = Alignment(horizontal="center", vertical="center")
            left        = Alignment(horizontal="left",   vertical="center",
                                    wrap_text=False)

            # ── Cabeçalho de metadados (linhas 1-2) ──────────────────────────
            metadata_rows = [
                [("DATA DE MAPEAMENTO", meta["data"]),
                 ("RESPONSÁVEL",        meta["responsavel"]),
                 ("BAUDRATE",           meta["baudrate"])],
                [("MODELO",             meta["modelo"]),
                 ("FABRICANTE",         meta["fabricante"]),
                 ("ANO",                meta["ano"])],
            ]
            # Escreve os metadados em 2 linhas, 3 pares (rótulo, valor) por linha.
            # Cada par ocupa 3 colunas: rótulo (1) + valor mesclado em 2 colunas.
            for r, items in enumerate(metadata_rows, start=1):
                col = 1
                for label, value in items:
                    # Coluna do rótulo (cinza, negrito)
                    lbl_cell = ws.cell(row=r, column=col, value=label)
                    lbl_cell.fill, lbl_cell.font = header_fill, bold
                    lbl_cell.border, lbl_cell.alignment = border, center
                    # Coluna do valor — mescla 2 colunas para ficar amplo
                    val_cell = ws.cell(row=r, column=col + 1, value=value)
                    val_cell.border, val_cell.alignment = border, left
                    # Aplica borda na segunda célula do merge (para a borda direita aparecer)
                    right_cell = ws.cell(row=r, column=col + 2)
                    right_cell.border = border
                    ws.merge_cells(start_row=r, start_column=col + 1,
                                   end_row=r,   end_column=col + 2)
                    col += 3   # avança para o próximo trio (rótulo+valor mesclado)

            # ── Cabeçalho da tabela (linha 4) ────────────────────────────────
            # Colunas da tabela VIRLOC. "MAT" e "VS" recebem as strings geradas
            # por _build_mat_string e _build_vs_string.
            headers = ["Prioridade", "CT", "INFORMAÇÕES", "PGN", "BYTE",
                       "MAT", "UNIDADE MEDIDA", "VS", "OBSERVAÇÃO"]
            HEADER_ROW = 4   # a tabela começa na linha 4 (1-2 = metadados, 3 = espaço)
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=HEADER_ROW, column=c, value=h)
                cell.fill      = header_fill
                cell.font      = bold
                cell.border    = border
                cell.alignment = center

            # Larguras das colunas (em "caracteres"). A coluna VS (8ª) é a mais
            # larga porque a string VS é longa.
            widths = [12, 6, 28, 8, 10, 26, 16, 70, 22]
            for c, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(c)].width = w

            # ── Validação preliminar ─────────────────────────────────────────
            # O VIRLOC só tem 25 filtros (VS1900..VS1924). Acima disso, avisa o
            # usuário que os sinais excedentes sairão com VS inválido.
            if len(self._mappings) > 25:
                resp = QMessageBox.question(
                    self, "Excesso de filtros",
                    f"Você tem {len(self._mappings)} sinais mapeados, mas o sistema "
                    f"VIRLOC suporta no máximo 25 filtros (VS1900..VS1924).\n\n"
                    "Os sinais excedentes serão exportados com VS inválido.\n"
                    "Deseja continuar mesmo assim?",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if resp != QMessageBox.Yes:
                    return

            # ── Linhas de dados ──────────────────────────────────────────────
            # Uma linha por sinal MAPEADO. Os contadores avançam juntos:
            row    = HEADER_ROW + 1   # primeira linha de dados
            ct     = 0           # CT sequencial — coluna "CT" (000, 001, ...)
            vs_fn  = 0           # número do filtro S19 (00..24) → cabeçalho VS19xx
            ct_num = 1           # número do CT no VS (01..96)
            for key, sig in self._mappings.items():
                nome    = _SIGNAL_INFO_NAME.get(key, key)      # nome amigável
                unidade = _SIGNAL_UNIT.get(key, "")            # unidade de medida
                vs_str  = _build_vs_string(vs_fn, ct_num, sig) # string VS do filtro
                mat_str = _build_mat_string(vs_fn, sig)        # string MAT (conversão)
                byte_str = str(sig.byte_index + 1)             # byte 1-indexed para a planilha
                if sig.length_bytes > 1:
                    byte_str += f"–{sig.byte_index + sig.length_bytes}"   # faixa se >1 byte

                # Sinais mapeados têm prioridade "Alta".
                values = ["Alta", f"{ct:03d}", nome,
                          str(sig.pgn) if sig.pgn else "",
                          byte_str, mat_str, unidade, vs_str, ""]
                fill = row_fill_a if ct % 2 == 0 else row_fill_b   # cor alternada (zebra)
                for c, v in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border    = border
                    cell.fill      = fill
                    # Colunas 3 (INFORMAÇÕES), 6 (MAT) e 8 (VS) alinham à esquerda.
                    cell.alignment = left if c in (3, 6, 8) else center
                row    += 1
                ct     += 1
                vs_fn  += 1   # próximo filtro VS
                ct_num += 1   # próximo CT

            # ── Sinais ainda não mapeados (linhas em branco com prioridade Baixa) ─
            # Lista TODOS os sinais do catálogo que ainda não foram identificados,
            # como linhas vazias de prioridade "Baixa" — servem de checklist do
            # que falta mapear, mantendo VS/MAT em branco.
            for key in TESTS:
                if key in self._mappings:
                    continue   # já mapeado, pula
                nome    = _SIGNAL_INFO_NAME.get(key, key)
                unidade = _SIGNAL_UNIT.get(key, "")
                values = ["Baixa", f"{ct:03d}", nome, "", "", "", unidade, "", ""]
                fill = row_fill_a if ct % 2 == 0 else row_fill_b
                for c, v in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border, cell.fill = border, fill
                    cell.alignment = left if c in (3, 6, 8) else center
                row += 1
                ct  += 1

            # ── Congelar painel + filtro automático ──────────────────────────
            # Congela tudo acima da primeira linha de dados (cabeçalho fica fixo
            # ao rolar) e ativa o auto-filtro do Excel sobre a faixa da tabela.
            ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)
            ws.auto_filter.ref = f"A{HEADER_ROW}:I{row - 1}"

            wb.save(path)   # grava o arquivo .xlsx em disco
            QMessageBox.information(
                self, "Exportado com sucesso",
                f"Mapeamento salvo em:\n{path}\n\n"
                f"{len(self._mappings)} sinal(is) exportado(s)."
            )
        except PermissionError:
            QMessageBox.critical(
                self, "Arquivo em uso",
                "Não foi possível salvar.\nFeche o arquivo no Excel e tente novamente."
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro ao exportar", str(e))

    @pyqtSlot()
    def _calibrate_selected(self):
        """Abre o assistente de calibração para o sinal selecionado na tabela.

        Valida pré-condições (há sinais, há linha selecionada, há conexão CAN),
        abre o _CalibrationDialog e, se o usuário aplicar, grava o novo
        fator/offset no CandidateSignal, reconstrói a string da fórmula e marca
        o sinal como calibrado manualmente (confiança ~99%).
        """
        if not self._mappings:
            QMessageBox.warning(self, "Sem sinais",
                                "Mapeie ao menos um sinal antes de calibrar.")
            return
        # currentRow() devolve o índice da linha selecionada; mapeamos de volta
        # para a chave do sinal via _row_keys (mesma ordem da tabela).
        row = self._table.currentRow()
        if row < 0 or row >= len(self._row_keys):
            QMessageBox.information(
                self, "Selecione um sinal",
                "Clique em uma linha da tabela para escolher o sinal a calibrar.")
            return
        key = self._row_keys[row]
        sig = self._mappings.get(key)
        if sig is None:
            return
        # A calibração lê valores AO VIVO; exige barramento conectado (ou replay).
        if self._bus is None or not self._bus.is_connected:
            QMessageBox.warning(
                self, "Sem conexão CAN",
                "A calibração precisa de dados ao vivo.\n"
                "Conecte ao barramento (ou inicie um replay CSV) antes.")
            return

        test = TESTS.get(key)
        signal_name = test.name if test else key
        # Abre o assistente (modal). exec_() retorna Accepted se o usuário aplicou.
        dlg = _CalibrationDialog(self._bus, sig, signal_name, self)
        if dlg.exec_() == QDialog.Accepted:
            factor, offset = dlg.result_factor_offset()   # (fator, offset) calculados
            if factor is not None:
                sig.formula_factor = factor               # grava no sinal
                sig.formula_offset = offset
                unit = _SIGNAL_UNIT.get(key, "")
                # Reconstrói o texto da fórmula. _fmt mostra inteiros sem casas
                # decimais e demais valores com até 6 dígitos significativos.
                def _fmt(v):
                    return str(int(v)) if v == int(v) else f"{v:.6g}"
                if offset == 0:
                    sig.formula_str = f"valor = raw × {_fmt(factor)}  [{unit}]"
                else:
                    s = "+" if offset >= 0 else "-"
                    sig.formula_str = (f"valor = raw × {_fmt(factor)} "
                                       f"{s} {_fmt(abs(offset))}  [{unit}]")
                # Calibração manual é confiável → eleva a confiança a (no mínimo) 99%.
                sig.confidence = max(sig.confidence, 0.99)
                self._rebuild_table()   # reflete a nova fórmula/confiança na tabela
                QMessageBox.information(
                    self, "Calibrado",
                    f"Sinal '{signal_name}' calibrado:\n\n{sig.formula_str}")

    @pyqtSlot()
    def _export_pdf_report(self):
        """Gera relatório PDF da sessão de mapeamento com avaliação.

        TÉCNICA: o relatório é montado como uma string HTML (_build_report_html)
        e renderizado em PDF pelo motor de impressão do Qt — um QTextDocument
        carrega o HTML e é "impresso" num QPrinter configurado para saída PDF.
        Não há dependência externa de PDF; tudo vem do próprio PyQt5.
        """
        if not self._mappings:
            QMessageBox.warning(self, "Nada para relatar",
                                "Nenhum sinal mapeado ainda.")
            return

        # 1. Coleta metadados do veículo
        dlg = _MetadataDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return
        meta = dlg.values()

        # 2. Escolhe local de salvamento
        modelo_clean = meta["modelo"].replace(" ", "_") or "veiculo"
        default_name = f"relatorio_{modelo_clean}_{meta['ano']}.pdf"
        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Relatório PDF", default_name, "PDF (*.pdf)"
        )
        if not path:
            return

        # 3. Gera o conteúdo HTML do relatório e o imprime para PDF.
        try:
            html = self._build_report_html(meta)   # monta o HTML completo
            doc = QTextDocument()
            # Registra os gráficos (QImage) como recursos ANTES de carregar o
            # HTML, para que as tags <img src="chart_<key>"> sejam resolvidas.
            self._register_chart_images(doc)
            doc.setHtml(html)                        # carrega o HTML no documento

            # Configura o "impressora virtual" para gerar um arquivo PDF A4.
            printer = QPrinter(QPrinter.HighResolution)   # alta resolução = texto nítido
            printer.setOutputFormat(QPrinter.PdfFormat)   # saída em PDF (não papel)
            printer.setOutputFileName(path)               # arquivo de destino
            printer.setPageSize(QPrinter.A4)
            # margens em mm (esquerda, topo, direita, base)
            printer.setPageMargins(15, 15, 15, 15, QPrinter.Millimeter)
            doc.print_(printer)                           # renderiza o HTML no PDF

            QMessageBox.information(
                self, "Relatório gerado",
                f"Relatório PDF salvo em:\n{path}\n\n"
                f"{len(self._mappings)} sinal(is) mapeado(s)."
            )
        except PermissionError:
            QMessageBox.critical(
                self, "Arquivo em uso",
                "Não foi possível salvar.\nFeche o PDF e tente novamente."
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro ao gerar PDF", str(e))

    def _build_report_html(self, meta: dict) -> str:
        """Monta o HTML do relatório com avaliação da sessão.

        Calcula estatísticas (cobertura, confiança média, distribuição), emite
        um veredito qualitativo, monta as tabelas de sinais mapeados/pendentes,
        a seção de candidatos alternativos e sugestões — tudo concatenado numa
        grande string HTML estilizada com CSS inline.
        """
        # ── Estatísticas ──────────────────────────────────────────────────────
        total_catalog = len(TESTS)              # total de sinais no catálogo
        total_mapped  = len(self._mappings)     # quantos mapeamos
        pct_mapped    = (total_mapped / total_catalog * 100) if total_catalog else 0  # % de cobertura
        avg_conf      = (sum(s.confidence for s in self._mappings.values())
                         / total_mapped) if total_mapped else 0   # confiança média
        # Contagem por faixa de confiança (alta / média / baixa):
        high_conf  = sum(1 for s in self._mappings.values() if s.confidence > 0.80)
        med_conf   = sum(1 for s in self._mappings.values() if 0.50 < s.confidence <= 0.80)
        low_conf   = sum(1 for s in self._mappings.values() if s.confidence <= 0.50)

        # Duração da sessão = agora - início (registrado na 1ª descoberta).
        if self._session_start:
            duration = datetime.now() - self._session_start
            mins = int(duration.total_seconds() // 60)
            secs = int(duration.total_seconds() % 60)
            duration_str = f"{mins} min {secs} s" if mins else f"{secs} s"
        else:
            duration_str = "—"

        # Avaliação qualitativa: traduz o % de cobertura em um veredito com cor.
        if pct_mapped >= 80:
            verdict, verdict_color = "✓ EXCELENTE — Sessão muito produtiva", "#2e7d32"
        elif pct_mapped >= 60:
            verdict, verdict_color = "✓ BOM — Sessão produtiva", "#558b2f"
        elif pct_mapped >= 40:
            verdict, verdict_color = "○ REGULAR — Sessão parcialmente produtiva", "#f57c00"
        elif pct_mapped >= 20:
            verdict, verdict_color = "△ BAIXO — Poucos sinais identificados", "#e65100"
        else:
            verdict, verdict_color = "✗ INSUFICIENTE — Refazer testes recomendado", "#c62828"

        # Conjunto de PGNs distintos validados (só faz sentido em J1939).
        pgns_validated = {sig.pgn for sig in self._mappings.values() if sig.pgn}
        pgns_str = ", ".join(str(p) for p in sorted(pgns_validated)) if pgns_validated else "Nenhum (protocolo proprietário ou 11-bit)"

        # Detecta se a sessão é majoritariamente 11-bit (proprietário) ou 29-bit (J1939).
        n_11bit = sum(1 for s in self._mappings.values() if not s.is_extended)
        n_29bit = total_mapped - n_11bit
        if n_29bit > n_11bit:
            protocol_desc = f"J1939 (29-bit) — {n_29bit} sinais 29-bit, {n_11bit} sinais 11-bit"
        elif n_11bit > 0:
            protocol_desc = f"Proprietário 11-bit — {n_11bit} sinais 11-bit, {n_29bit} sinais 29-bit"
        else:
            protocol_desc = "Indeterminado"

        # ── Tabela de sinais mapeados ─────────────────────────────────────────
        # Monta as linhas <tr> da tabela de mapeados, uma por sinal, com ícone e
        # cor de confiança (verde/laranja/vermelho).
        rows_mapped = ""
        for key, sig in self._mappings.items():
            test = TESTS.get(key)
            nome = test.name if test else key
            id_str = f"0x{sig.can_id:08X}" if sig.is_extended else f"0x{sig.can_id:04X}"
            byte_s = f"{sig.byte_index}" if sig.length_bytes == 1 else f"{sig.byte_index}-{sig.byte_index + sig.length_bytes - 1}"
            pgn_s = str(sig.pgn) if sig.pgn else "—"
            conf_pct = sig.confidence * 100
            if sig.confidence > 0.80:
                conf_color = "#2e7d32"
                conf_icon  = "✓"
            elif sig.confidence > 0.50:
                conf_color = "#f57c00"
                conf_icon  = "○"
            else:
                conf_color = "#c62828"
                conf_icon  = "△"

            rows_mapped += f"""
            <tr>
                <td>{nome}</td>
                <td style="font-family: Consolas, monospace;">{id_str}</td>
                <td align="center">{pgn_s}</td>
                <td align="center">{byte_s}</td>
                <td style="font-family: Consolas, monospace; font-size: 10px;">{sig.formula_str}</td>
                <td align="center" style="color: {conf_color}; font-weight: bold;">{conf_icon} {conf_pct:.0f}%</td>
            </tr>
            """

        # ── Sinais NÃO mapeados ───────────────────────────────────────────────
        # Linhas da tabela de pendentes: tudo do catálogo que não foi mapeado.
        rows_unmapped = ""
        unmapped_keys = [k for k in TESTS if k not in self._mappings]
        for key in unmapped_keys:
            test = TESTS[key]
            rows_unmapped += f"""
            <tr>
                <td>{test.name}</td>
                <td>{test.unit or '—'}</td>
                <td style="color: #c62828;">Não identificado</td>
            </tr>
            """

        # ── Sugestões de melhoria ─────────────────────────────────────────────
        # Gera recomendações automáticas conforme o que foi observado na sessão
        # (sinais de baixa/média confiança, pendentes, protocolo proprietário).
        suggestions = []
        if low_conf > 0:
            suggestions.append(
                f"<b>{low_conf} sinal(is) com baixa confiança (&lt;50%):</b> "
                f"Recomendamos refazer estes testes com movimento mais lento e amplo. "
                f"O algoritmo pode ter confundido com bytes de ruído.")
        if med_conf > 0:
            suggestions.append(
                f"<b>{med_conf} sinal(is) com confiança média (50-80%):</b> "
                f"Vale validar manualmente no monitor CAN se o valor decodificado "
                f"bate com o instrumento do painel.")
        if unmapped_keys:
            unmapped_names = ", ".join(TESTS[k].name for k in unmapped_keys[:5])
            extra = f" e mais {len(unmapped_keys)-5}" if len(unmapped_keys) > 5 else ""
            suggestions.append(
                f"<b>Sinais ainda não mapeados:</b> {unmapped_names}{extra}.<br>"
                f"Considere refazer os testes para esses sinais ou validar se "
                f"o veículo realmente os disponibiliza no barramento.")
        if n_11bit > 0 and not pgns_validated:
            suggestions.append(
                "<b>Veículo aparentemente usa protocolo proprietário (11-bit):</b> "
                "validação manual de cada sinal é fortemente recomendada antes de "
                "implantar no sistema de telemetria.")
        if not suggestions:
            suggestions.append("Nenhuma — sessão completa e satisfatória. ✓")

        # Converte a lista de sugestões em uma lista HTML <ul><li>...</li></ul>.
        suggestions_html = "<ul>" + "".join(f"<li>{s}</li>" for s in suggestions) + "</ul>"

        # ── Montagem do HTML ──────────────────────────────────────────────────
        # Concatena todas as seções num único documento HTML estilizado. As
        # f-strings interpolam as variáveis calculadas acima. Os comentários
        # <!-- ... --> abaixo apenas delimitam cada seção visualmente.
        now_str = datetime.now().strftime("%d/%m/%Y às %H:%M")   # data/hora de geração

        html = f"""
        <html><head><meta charset="utf-8"></head>
        <body style="font-family: Arial, sans-serif; color: #222;">

        <h1 style="color: #1a237e; border-bottom: 3px solid #1a237e; padding-bottom: 8px;">
            Relatório de Mapeamento CAN
        </h1>
        <p style="color: #666; margin: 0;">Gerado em {now_str} pelo IxxatInterface v7</p>

        <!-- METADADOS DO VEÍCULO -->
        <h2 style="color: #1a237e; margin-top: 24px;">📋 Identificação do Veículo</h2>
        <table cellpadding="6" cellspacing="0" border="0" width="100%" style="background:#f5f5f5;">
            <tr><td width="180"><b>Data do mapeamento:</b></td><td>{meta['data']}</td>
                <td width="140"><b>Responsável:</b></td><td>{meta['responsavel'] or '—'}</td></tr>
            <tr><td><b>Modelo:</b></td><td>{meta['modelo'] or '—'}</td>
                <td><b>Fabricante:</b></td><td>{meta['fabricante'] or '—'}</td></tr>
            <tr><td><b>Ano:</b></td><td>{meta['ano'] or '—'}</td>
                <td><b>Baudrate CAN:</b></td><td>{meta['baudrate']} kbps</td></tr>
        </table>

        <!-- RESUMO DA SESSÃO -->
        <h2 style="color: #1a237e; margin-top: 24px;">📊 Resumo da Sessão</h2>
        <table cellpadding="6" cellspacing="0" border="0" width="100%" style="background:#f5f5f5;">
            <tr><td width="220"><b>Tempo total de análise:</b></td><td>{duration_str}</td></tr>
            <tr><td><b>Sinais no catálogo:</b></td><td>{total_catalog}</td></tr>
            <tr><td><b>Sinais mapeados:</b></td><td><b>{total_mapped}</b> ({pct_mapped:.0f}%)</td></tr>
            <tr><td><b>Confiança média:</b></td><td>{avg_conf * 100:.0f}%</td></tr>
            <tr><td><b>Distribuição de confiança:</b></td>
                <td>
                    <span style="color:#2e7d32;">✓ Alta (&gt;80%): {high_conf}</span> &nbsp;
                    <span style="color:#f57c00;">○ Média (50-80%): {med_conf}</span> &nbsp;
                    <span style="color:#c62828;">△ Baixa (≤50%): {low_conf}</span>
                </td></tr>
            <tr><td><b>Protocolo dominante:</b></td><td>{protocol_desc}</td></tr>
            <tr><td><b>PGNs J1939 validadas:</b></td><td style="font-family: Consolas, monospace;">{pgns_str}</td></tr>
        </table>

        <!-- VEREDITO -->
        <div style="margin-top:24px; padding: 16px; background:{verdict_color}; color:white; font-size: 16px; font-weight: bold; text-align: center;">
            {verdict} ({pct_mapped:.0f}% de aproveitamento)
        </div>

        <!-- TABELA SINAIS MAPEADOS -->
        <h2 style="color: #1a237e; margin-top: 24px;">✓ Sinais Mapeados ({total_mapped})</h2>
        <table cellpadding="6" cellspacing="0" border="1" bordercolor="#bbb" width="100%" style="border-collapse: collapse; font-size: 11px;">
            <thead style="background:#1a237e; color:white;">
                <tr>
                    <th>Sinal</th>
                    <th>CAN ID</th>
                    <th>PGN</th>
                    <th>Byte(s)</th>
                    <th>Fórmula</th>
                    <th>Confiança</th>
                </tr>
            </thead>
            <tbody>{rows_mapped}</tbody>
        </table>

        <!-- SINAIS NÃO MAPEADOS -->
        <h2 style="color: #1a237e; margin-top: 24px;">✗ Sinais Pendentes ({len(unmapped_keys)})</h2>
        {"<table cellpadding='6' cellspacing='0' border='1' bordercolor='#bbb' width='100%' style='border-collapse: collapse; font-size: 11px;'><thead style='background:#666; color:white;'><tr><th>Sinal</th><th>Unidade</th><th>Status</th></tr></thead><tbody>" + rows_unmapped + "</tbody></table>" if unmapped_keys else "<p style='color:#2e7d32;'>Todos os sinais do catálogo foram mapeados! ✓</p>"}

        <!-- GRÁFICOS: COMPORTAMENTO OBSERVADO (LINHA) + PROBABILIDADE (BARRAS) -->
        <h2 style="color: #1a237e; margin-top: 24px;">📈 Gráficos — Comportamento Observado e Probabilidade</h2>
        <p style="color: #555; font-size: 11px;">
        Para cada sinal testado, o gráfico mostra <b>todas as PGNs/CAN IDs que o algoritmo
        avaliou</b> e a <b>confiança</b> (probabilidade) de cada uma. Barras verdes = alta
        probabilidade; laranja = média; vermelha = baixa.
        </p>
        {self._build_charts_html()}

        <!-- CANDIDATOS ALTERNATIVOS (PGNs menos prováveis) -->
        <h2 style="color: #1a237e; margin-top: 24px;">🔍 Candidatos Alternativos (PGNs menos prováveis)</h2>
        <p style="color: #555; font-size: 11px;">
        Lista de candidatos que o algoritmo identificou mas <b>não foram escolhidos como melhor opção</b>.
        Útil para investigação futura caso o sinal mapeado se mostre incorreto durante a validação.
        </p>
        {self._build_alternatives_html()}

        <!-- RECOMENDAÇÕES -->
        <h2 style="color: #1a237e; margin-top: 24px;">💡 Avaliação e Recomendações</h2>
        {suggestions_html}

        <hr style="margin-top: 30px; border: none; border-top: 1px solid #ccc;">
        <p style="color: #888; font-size: 10px; text-align: center;">
            Relatório gerado automaticamente pelo IxxatInterface v7 — Ferramenta de mapeamento de sinais CAN<br>
            Documento técnico interno — uso restrito da equipe de engenharia
        </p>
        </body></html>
        """
        return html

    def _build_charts_html(self) -> str:
        """Monta a seção de gráficos (um <img> por sinal testado).

        Cada imagem é resolvida em tempo de impressão a partir dos recursos
        registrados em _register_chart_images(doc). Aqui só emitimos as tags
        <img src="chart_<key>"> na ordem dos sinais com candidatos.
        """
        if not self._all_candidates:
            return "<p style='color:#888;'>Nenhuma PGN foi checada nesta sessão.</p>"
        blocks = []
        for key, candidates in self._all_candidates.items():
            if not candidates:
                continue
            test = TESTS.get(key)
            nome = test.name if test else key
            block = f'<h3 style="color:#1a237e; margin-top:18px; margin-bottom:2px;">{nome}</h3>'
            # 1º: gráfico de LINHA do comportamento observado (se houver amostras)
            top = candidates[0]
            if getattr(top, "samples", None):
                block += (
                    '<div style="margin-top:4px;">'
                    '<span style="color:#555; font-size:11px;">Comportamento observado durante o teste:</span><br>'
                    f'<img src="behavior_{key}" width="760">'
                    '</div>'
                )
            # 2º: gráfico de barras de probabilidade dos candidatos
            block += (
                '<div style="margin-top:6px;">'
                '<span style="color:#555; font-size:11px;">Probabilidade das PGNs checadas:</span><br>'
                f'<img src="chart_{key}" width="760">'
                '</div>'
            )
            blocks.append(block)
        if not blocks:
            return "<p style='color:#888;'>Nenhuma PGN foi checada nesta sessão.</p>"
        return "".join(blocks)

    def _register_chart_images(self, doc):
        """Desenha e registra no QTextDocument um gráfico por sinal testado.

        Cada gráfico é uma QImage (barras de confiança dos candidatos) registrada
        como recurso de imagem do documento, referenciada no HTML por
        'chart_<key>'. Deve ser chamado ANTES de doc.print_().
        """
        from PyQt5.QtCore import QUrl
        for key, candidates in self._all_candidates.items():
            if not candidates:
                continue
            test = TESTS.get(key)
            nome = test.name if test else key
            unit = _SIGNAL_UNIT.get(key, "")
            # Gráfico de barras de confiança (probabilidade dos candidatos)
            img_bar = self._make_candidates_chart(nome, candidates)
            doc.addResource(QTextDocument.ImageResource,
                            QUrl(f"chart_{key}"), img_bar)
            # Gráfico de LINHA do comportamento observado (forma de onda do
            # melhor candidato durante o teste). Só registra se houver amostras.
            top = candidates[0]
            if getattr(top, "samples", None):
                img_line = self._make_behavior_chart(nome, top, unit)
                doc.addResource(QTextDocument.ImageResource,
                                QUrl(f"behavior_{key}"), img_line)

    def _make_candidates_chart(self, nome: str, candidates: list):
        """Desenha um gráfico de barras horizontais da confiança dos candidatos.

        Cada barra = um candidato (PGN/CAN ID + byte). O comprimento da barra é
        proporcional à confiança (0–100%). Cor por faixa: verde >80%, laranja
        50–80%, vermelho <50%. Retorna uma QImage pronta para embutir no PDF.
        """
        from PyQt5.QtGui import QImage, QPainter, QPen, QFont
        from PyQt5.QtCore import QRectF, Qt as _Qt

        n = min(len(candidates), 8)        # no máximo 8 barras por gráfico
        row_h  = 28                         # altura de cada linha/barra
        top    = 34                         # espaço do título no topo
        W      = 760
        H      = top + n * row_h + 14
        label_w = 250                       # largura reservada ao rótulo (PGN/ID)
        bar_x   = label_w + 10
        bar_max = W - bar_x - 70            # largura máxima da barra (sobra p/ %)

        img = QImage(W, H, QImage.Format_ARGB32)
        img.fill(QColor("#ffffff"))
        p = QPainter(img)
        p.setRenderHint(QPainter.Antialiasing)

        # Título do gráfico (nome do sinal testado)
        p.setPen(QColor("#1a237e"))
        p.setFont(QFont("Arial", 11, QFont.Bold))
        p.drawText(QRectF(4, 6, W - 8, 22), _Qt.AlignLeft | _Qt.AlignVCenter, nome)

        for i, sig in enumerate(candidates[:n]):
            y = top + i * row_h
            conf = max(0.0, min(1.0, sig.confidence))

            # Rótulo: PGN/sigla + CAN ID + byte
            id_str = (f"0x{sig.can_id:08X}" if sig.is_extended
                      else f"0x{sig.can_id:04X}")
            byte_s = (f"B{sig.byte_index}" if sig.length_bytes == 1
                      else f"B{sig.byte_index}-{sig.byte_index + sig.length_bytes - 1}")
            pgn_lbl = sig.pgn_name or (str(sig.pgn) if sig.pgn else id_str)
            label = f"{pgn_lbl}  {id_str} {byte_s}"
            p.setPen(QColor("#333333"))
            p.setFont(QFont("Consolas", 8))
            p.drawText(QRectF(4, y, label_w, row_h - 4),
                       _Qt.AlignLeft | _Qt.AlignVCenter, label)

            # Trilho de fundo da barra
            p.fillRect(QRectF(bar_x, y + 5, bar_max, row_h - 14), QColor("#eeeeee"))
            # Cor por faixa de confiança
            if conf > 0.80:
                color = QColor("#2e7d32")    # verde — alta
            elif conf > 0.50:
                color = QColor("#f57c00")    # laranja — média
            else:
                color = QColor("#c62828")    # vermelho — baixa
            # Barra preenchida proporcional à confiança
            p.fillRect(QRectF(bar_x, y + 5, bar_max * conf, row_h - 14), color)

            # Texto da porcentagem ao lado da barra
            p.setPen(QColor("#000000"))
            p.setFont(QFont("Arial", 9, QFont.Bold))
            p.drawText(QRectF(bar_x + bar_max + 6, y, 60, row_h - 4),
                       _Qt.AlignLeft | _Qt.AlignVCenter, f"{conf * 100:.0f}%")

        p.end()
        return img

    def _make_behavior_chart(self, nome: str, sig, unit: str):
        """Desenha um GRÁFICO DE LINHA do comportamento observado do sinal.

        Plota a forma de onda (sig.samples — valores em engenharia capturados
        durante o teste) ao longo do tempo. Mostra como o sinal de fato variou
        (ex.: RPM subindo 800→1500→2500→3500). Eixo Y rotulado com mín/máx.
        Retorna uma QImage pronta para embutir no PDF.
        """
        from PyQt5.QtGui import QImage, QPainter, QPen, QFont
        from PyQt5.QtCore import QRectF, QPointF, Qt as _Qt

        samples = list(sig.samples or [])
        W, H = 760, 200
        img = QImage(W, H, QImage.Format_ARGB32)
        img.fill(QColor("#ffffff"))
        p = QPainter(img)
        p.setRenderHint(QPainter.Antialiasing)

        # Margens da área de plotagem
        MX, MY = 70, 30      # esquerda (rótulos Y) / topo (título)
        MR, MB = 16, 26      # direita / base (eixo X)
        PW = W - MX - MR
        PH = H - MY - MB

        # Título
        id_str = (f"0x{sig.can_id:08X}" if sig.is_extended else f"0x{sig.can_id:04X}")
        byte_s = (f"B{sig.byte_index}" if sig.length_bytes == 1
                  else f"B{sig.byte_index}-{sig.byte_index + sig.length_bytes - 1}")
        titulo = f"{nome}  ·  {id_str} {byte_s}"
        p.setPen(QColor("#1a237e"))
        p.setFont(QFont("Arial", 11, QFont.Bold))
        p.drawText(QRectF(4, 4, W - 8, 20), _Qt.AlignLeft | _Qt.AlignVCenter, titulo)

        # Moldura da área de plotagem
        p.setPen(QPen(QColor("#cccccc"), 1))
        p.drawRect(QRectF(MX, MY, PW, PH))

        if len(samples) < 2:
            p.setPen(QColor("#999999"))
            p.setFont(QFont("Arial", 10))
            p.drawText(QRectF(MX, MY, PW, PH), _Qt.AlignCenter,
                       "Sem amostras suficientes para o gráfico")
            p.end()
            return img

        vmin, vmax = min(samples), max(samples)
        if vmax == vmin:                  # série constante: evita divisão por zero
            vmax = vmin + 1.0
        span = vmax - vmin

        # Rótulos do eixo Y (mín, meio, máx) + linhas de grade
        p.setFont(QFont("Consolas", 8))
        for frac, val in ((0.0, vmax), (0.5, (vmin + vmax) / 2), (1.0, vmin)):
            gy = MY + frac * PH
            p.setPen(QPen(QColor("#eeeeee"), 1))
            p.drawLine(QPointF(MX, gy), QPointF(MX + PW, gy))
            p.setPen(QColor("#666666"))
            lbl = f"{val:.1f}" if span < 100 else f"{val:.0f}"
            p.drawText(QRectF(2, gy - 8, MX - 6, 16),
                       _Qt.AlignRight | _Qt.AlignVCenter, lbl)

        # Linha do comportamento observado
        p.setPen(QPen(QColor("#6c63ff"), 2))
        n = len(samples)
        pts = []
        for i, v in enumerate(samples):
            x = MX + (i / (n - 1)) * PW
            y = MY + (1.0 - (v - vmin) / span) * PH
            pts.append(QPointF(x, y))
        for i in range(1, len(pts)):
            p.drawLine(pts[i - 1], pts[i])

        # Eixo X (rótulo de tempo)
        p.setPen(QColor("#666666"))
        p.setFont(QFont("Arial", 8))
        p.drawText(QRectF(MX, MY + PH + 4, PW, 18),
                   _Qt.AlignRight | _Qt.AlignVCenter,
                   f"tempo do teste →   (unidade: {unit or 'bruto'})")

        p.end()
        return img

    def _build_alternatives_html(self) -> str:
        """Monta a seção de candidatos alternativos do relatório PDF.

        Para cada sinal que teve mais de um candidato, lista do 2º ao 9º (o 1º
        já aparece em "Sinais Mapeados"). Serve de pista para investigação caso
        o sinal escolhido se mostre incorreto durante a validação.
        """
        if not self._all_candidates:
            return "<p style='color:#888;'>Nenhuma alternativa registrada nesta sessão.</p>"

        blocks = []
        for key, candidates in self._all_candidates.items():
            if len(candidates) <= 1:
                continue   # sem alternativas — só houve o top 1
            test = TESTS.get(key)
            nome = test.name if test else key
            # Pula o primeiro (já está em "Sinais Mapeados") e mostra os outros.
            # candidates[1:9] = do 2º ao 9º; start=2 numera o rank a partir de #2.
            alt_rows = ""
            for i, sig in enumerate(candidates[1:9], start=2):
                id_str = (f"0x{sig.can_id:08X}" if sig.is_extended
                          else f"0x{sig.can_id:04X}")
                byte_s = (f"{sig.byte_index}" if sig.length_bytes == 1
                          else f"{sig.byte_index}-{sig.byte_index + sig.length_bytes - 1}")
                pgn_s = str(sig.pgn) if sig.pgn else "—"
                pgn_n = sig.pgn_name or "—"
                conf_pct = sig.confidence * 100
                if sig.confidence > 0.80:
                    conf_color = "#2e7d32"
                elif sig.confidence > 0.50:
                    conf_color = "#f57c00"
                else:
                    conf_color = "#c62828"
                alt_rows += f"""
                <tr>
                    <td align="center">#{i}</td>
                    <td style="font-family: Consolas, monospace;">{id_str}</td>
                    <td align="center">{pgn_s}</td>
                    <td align="center">{pgn_n}</td>
                    <td align="center">{byte_s}</td>
                    <td style="font-family: Consolas, monospace; font-size: 10px;">{sig.formula_str}</td>
                    <td align="center" style="color: {conf_color}; font-weight: bold;">{conf_pct:.0f}%</td>
                </tr>"""

            if not alt_rows:
                continue

            blocks.append(f"""
            <h3 style="color: #555; margin-top: 16px; margin-bottom: 4px;
                       background:#eee; padding: 6px 10px;">
                {nome}
                <span style="color:#888; font-weight: normal; font-size: 12px;">
                    ({len(candidates) - 1} candidato(s) alternativo(s))
                </span>
            </h3>
            <table cellpadding="5" cellspacing="0" border="1" bordercolor="#ccc"
                   width="100%" style="border-collapse: collapse; font-size: 10px;">
                <thead style="background:#aaa; color:white;">
                    <tr>
                        <th width="40">Rank</th>
                        <th>CAN ID</th>
                        <th>PGN</th>
                        <th>Nome PGN</th>
                        <th>Byte(s)</th>
                        <th>Fórmula</th>
                        <th width="80">Confiança</th>
                    </tr>
                </thead>
                <tbody>{alt_rows}</tbody>
            </table>""")

        if not blocks:
            return "<p style='color:#888;'>Nenhum sinal teve mais de uma alternativa registrada.</p>"
        return "".join(blocks)   # junta todos os blocos de alternativas

    @pyqtSlot()
    def _clear(self):
        """Limpa toda a sessão: mapeamentos, candidatos, tabela e resumo."""
        self._mappings.clear()           # remove os sinais mapeados
        self._all_candidates.clear()     # remove os candidatos guardados
        self._session_start = None       # zera o cronômetro da sessão
        self._table.setRowCount(0)       # esvazia a tabela
        self._lbl_summary.setText("Nenhum sinal mapeado ainda.")
